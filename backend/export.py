import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PASTEL_COLORS = [
    "FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF",
    "E8BAFF", "FFBAF3", "BAF7FF", "D4FFBA", "FFD9BA",
    "C9BAFF", "FFBAD9",
]

LIGHT_GRAY = "D3D3D3"
HEADER_GRAY = "A0A0A0"
ROW_HEIGHT = 60  # tall enough for course + instructor + optional day labels


def _faculty_color(faculty_id: int | None) -> str:
    if faculty_id is None:
        return "E0E0E0"
    return PASTEL_COLORS[faculty_id % len(PASTEL_COLORS)]


def _credit_hours(course_number: int) -> int:
    """2nd digit of the 4-digit course number encodes credit hours."""
    return (course_number // 100) % 10


def generate_excel(db, term_id: int) -> bytes:
    from models import Term, TimeSlot, Room, TaughtWithGroup

    term = db.query(Term).filter(Term.id == term_id).first()
    if not term:
        raise ValueError("Term not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    time_slots = db.query(TimeSlot).order_by(TimeSlot.display_order).all()
    # Online rooms sorted last
    rooms = sorted(
        db.query(Room).all(),
        key=lambda r: (r.is_online, r.label)
    )

    if not time_slots or not rooms:
        ws["A1"] = "No data"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    row = 1

    # Term header
    term_label = f"Term: {term.semester.name.value.capitalize()} {term.year}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(rooms) + 1)
    cell = ws.cell(row=row, column=1, value=term_label)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="right")
    row += 1

    for table in term.schedule_tables:
        row += 1  # empty row between tables

        # Weekday header row
        weekday_names = " / ".join(w.name.value.capitalize() for w in sorted(table.weekdays, key=lambda w: w.display_order))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(rooms) + 1)
        wd_cell = ws.cell(row=row, column=1, value=weekday_names)
        wd_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)
        wd_cell.font = Font(bold=True)
        wd_cell.alignment = Alignment(horizontal="right")
        row += 1

        # Column headers: Time Slot | Room1 | Room2 | ...
        ws.cell(row=row, column=1, value="Time Slot").font = Font(bold=True)
        for col_idx, room in enumerate(rooms, start=2):
            hdr = room.label if room.is_online else f"{room.label} ({room.capacity})"
            ws.cell(row=row, column=col_idx, value=hdr).font = Font(bold=True)
        row += 1

        table_start_row = row

        # Map time_slot_id -> row index within this table
        slot_row_map = {ts.id: table_start_row + i for i, ts in enumerate(time_slots)}

        # Set row heights and write time slot labels
        for i, ts in enumerate(time_slots):
            r = table_start_row + i
            ws.row_dimensions[r].height = ROW_HEIGHT
            ws.cell(row=r, column=1, value=ts.label)

        # Write entries
        room_col_map = {r.id: col_idx for col_idx, r in enumerate(rooms, start=2)}
        table_weekday_ids = {w.id for w in table.weekdays}

        for entry in table.entries:
            if not entry.room_id or not entry.time_slots:
                continue
            col = room_col_map.get(entry.room_id)
            if col is None:
                continue

            entry_slots = sorted(entry.time_slots, key=lambda ts: ts.display_order)
            if not entry_slots:
                continue

            first_slot = entry_slots[0]
            last_slot = entry_slots[-1]
            start_r = slot_row_map.get(first_slot.id)
            end_r = slot_row_map.get(last_slot.id)
            if start_r is None or end_r is None:
                continue

            course = entry.course
            text = f"{course.dept_code} {course.course_number}\n{course.course_name}"
            if entry.faculty:
                text += f"\n{entry.faculty.last_name}, {entry.faculty.first_name}"

            # Add active weekdays if this entry runs on a subset of the table's days
            if entry.active_weekdays:
                active_ids = {w.id for w in entry.active_weekdays}
                if active_ids < table_weekday_ids:
                    day_str = " ".join(
                        w.name.value.capitalize()[:2]
                        for w in sorted(entry.active_weekdays, key=lambda w: w.display_order)
                    )
                    text += f"\n[{day_str}]"

            fill_color = _faculty_color(entry.faculty_id)
            fill = PatternFill(fill_type="solid", fgColor=fill_color)

            if start_r == end_r:
                c = ws.cell(row=start_r, column=col, value=text)
                c.fill = fill
                c.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                ws.merge_cells(start_row=start_r, start_column=col, end_row=end_r, end_column=col)
                c = ws.cell(row=start_r, column=col, value=text)
                c.fill = fill
                c.alignment = Alignment(wrap_text=True, vertical="center")

        row = table_start_row + len(time_slots)

    # Auto-size columns
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, len(rooms) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # ── Faculty Load sheet ────────────────────────────────────────────────────
    ws_load = wb.create_sheet("Faculty Load")

    # Build TaughtWith lookups
    tw_groups = db.query(TaughtWithGroup).all()
    course_to_tw: dict[int, int] = {}
    tw_courses: dict[int, list[int]] = {}
    for g in tw_groups:
        ids = [m.course_id for m in g.members]
        tw_courses[g.id] = ids
        for cid in ids:
            course_to_tw[cid] = g.id

    # Collect scheduled entries with faculty
    all_entries = [
        e for e in term.schedule_entries
        if e.schedule_table_id and e.faculty_id
    ]

    by_faculty: dict[int, list] = {}
    for e in all_entries:
        by_faculty.setdefault(e.faculty_id, []).append(e)

    from models import LoadSettings
    settings = db.query(LoadSettings).first()
    fulltime_load = settings.fulltime_load if settings else 3
    parttime_load = settings.parttime_load if settings else 2

    def _full_load(f) -> int:
        return fulltime_load if f.rank.value == "full_time" else parttime_load

    load_rows: list[dict] = []
    for fid, fentries in by_faculty.items():
        faculty = fentries[0].faculty
        units: dict[str, list] = {}
        for e in fentries:
            gid = course_to_tw.get(e.course_id)
            key = f"tw_{gid}" if gid else f"c_{e.course_id}"
            units.setdefault(key, []).append(e)

        courses_data = []
        total_sections = 0
        total_ch = 0

        for key, uentries in units.items():
            if key.startswith("tw_"):
                gid = int(key[3:])
                seen: dict[int, object] = {}
                for e in uentries:
                    seen.setdefault(e.course_id, e.course)
                sorted_c = sorted(seen.values(), key=lambda c: c.course_number)
                display = " / ".join(f"{c.dept_code} {c.course_number} {c.course_name}" for c in sorted_c)
                ch = sum(_credit_hours(c.course_number) for c in sorted_c)
                rep_cid = next(iter(seen))
                sections = sum(1 for e in uentries if e.course_id == rep_cid)
            else:
                course = uentries[0].course
                display = f"{course.dept_code} {course.course_number} {course.course_name}"
                ch = _credit_hours(course.course_number)
                sections = len(uentries)

            total_sections += sections
            total_ch += ch * sections
            courses_data.append((display, sections, ch, ch * sections))

        courses_data.sort(key=lambda x: x[0])
        load_rows.append({
            "name": f"{faculty.last_name}, {faculty.first_name}",
            "rank": faculty.rank.value.replace("_", " ").title(),
            "full_load": _full_load(faculty),
            "courses": courses_data,
            "total_sections": total_sections,
            "total_credit_hours": total_ch,
        })

    load_rows.sort(key=lambda x: x["name"])

    # Write Faculty Load sheet
    hdr_fill = PatternFill(fill_type="solid", fgColor=HEADER_GRAY)
    hdr_font = Font(bold=True, color="FFFFFF")

    load_row = 1
    for faculty_data in load_rows:
        # Faculty name header
        ws_load.merge_cells(start_row=load_row, start_column=1, end_row=load_row, end_column=4)
        name_cell = ws_load.cell(row=load_row, column=1,
                                  value=f"{faculty_data['name']}  ({faculty_data['rank']} — load {faculty_data['full_load']})")
        name_cell.font = Font(bold=True, size=12)
        name_cell.fill = PatternFill(fill_type="solid", fgColor="C8D8E8")
        load_row += 1

        # Column headers
        for col, hdr in enumerate(["Course(s)", "Sections", "Credit Hrs / unit", "Total Credit Hrs"], start=1):
            c = ws_load.cell(row=load_row, column=col, value=hdr)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="right")
        load_row += 1

        # Course rows
        for display, sections, ch, total in faculty_data["courses"]:
            ws_load.cell(row=load_row, column=1, value=display)
            ws_load.cell(row=load_row, column=2, value=sections).alignment = Alignment(horizontal="right")
            ws_load.cell(row=load_row, column=3, value=ch).alignment = Alignment(horizontal="right")
            ws_load.cell(row=load_row, column=4, value=total).alignment = Alignment(horizontal="right")
            load_row += 1

        # Totals row
        tot_fill = PatternFill(fill_type="solid", fgColor="E8E8E8")
        for col in range(1, 5):
            ws_load.cell(row=load_row, column=col).fill = tot_fill
        ws_load.cell(row=load_row, column=1, value="TOTAL").font = Font(bold=True)
        ws_load.cell(row=load_row, column=2, value=faculty_data["total_sections"]).font = Font(bold=True)
        ws_load.cell(row=load_row, column=2).alignment = Alignment(horizontal="right")
        ws_load.cell(row=load_row, column=4, value=faculty_data["total_credit_hours"]).font = Font(bold=True)
        ws_load.cell(row=load_row, column=4).alignment = Alignment(horizontal="right")
        load_row += 2  # blank row between faculty

    ws_load.column_dimensions["A"].width = 55
    ws_load.column_dimensions["B"].width = 12
    ws_load.column_dimensions["C"].width = 20
    ws_load.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
