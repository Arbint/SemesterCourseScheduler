"""Add/Keep/Change/Delete change-list generation (feedback_42).

Diffs a registrar-provided draft spreadsheet (last year's schedule for the
department) against the live schedule of a chosen term, and can re-export
the annotated result in the same spreadsheet layout.
"""
import io
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from export import _credit_hours

DAY_LETTERS = {"mon": "M", "tue": "T", "wed": "W", "thu": "R", "fri": "F"}

HEADER = [
    "ADD/KEEP/CHANGE/DELETE", "Term", "Start Date", "End Date", "CRN", "Subject",
    "CRSE#", "SEC#", "Course Title", "Type", "Inst. Method", "Instructor",
    "Secondary Instructor", "Hours", "Enrollment Max", "Waitlist Cap", "Begin",
    "End", "Days", "Bldg", "RM", "Course Comments (Visible on Online Schedule)",
    "Prerequisite", "Fee Detail", "Fee Amount", "Signature Restriction Code",
    "Signature Restriction Required",
]

FIELD_ORDER = [
    "term_num", "start_date", "end_date", "crn", "subject", "course_number",
    "section", "course_title", "type", "inst_method", "instructor",
    "secondary_instructor", "hours", "enrollment_max", "waitlist_cap", "begin",
    "end", "days", "bldg", "rm", "course_comments", "prerequisite",
    "fee_detail", "fee_amount", "sig_code", "sig_required",
]

STATUS_LABEL = {"keep": "KEEP", "changed": "CHANGE", "delete": "DELETE", "add": "ADD"}

# Which display column(s) light up yellow for a given changed_fields entry.
CHANGED_FIELD_COLUMNS = {
    "days": ["days"],
    "begin": ["begin"],
    "end": ["end"],
    "room": ["bldg", "rm"],
    "instructor": ["instructor"],
    "course_comment": ["course_comments"],
    "enrollment_max": ["enrollment_max"],
}

# Best-effort extraction of a "DEPT NUMBER" course reference from old free-text
# Course Comments (e.g. "Taught W/ CRN 91910 ANGD 6211" -> ("ANGD", 6211)).
# CRNs in the real data are always 5 digits and course numbers are 3-4, so
# requiring 3-4 digits naturally skips over the "CRN <number>" token.
_TAUGHTWITH_REF_RE = re.compile(r"(?<!\d)([A-Z]{2,10}(?:-[A-Z]+)?)\s+(\d{3,4})(?!\d)")


def extract_taughtwith_ref(comment):
    if not comment:
        return None
    for m in _TAUGHTWITH_REF_RE.finditer(comment):
        dept, num = m.group(1), m.group(2)
        if dept == "CRN":
            continue
        return (dept, int(num))
    return None


def time_to_hhmm(time_str):
    if not time_str:
        return None
    return int(time_str.replace(":", ""))


def entry_days(entry, table):
    table_weekdays = sorted(table.weekdays, key=lambda w: w.display_order)
    if entry.active_weekdays:
        active_ids = {w.id for w in entry.active_weekdays}
        table_ids = {w.id for w in table.weekdays}
        if active_ids and active_ids <= table_ids:
            return sorted(entry.active_weekdays, key=lambda w: w.display_order)
    return table_weekdays


def days_string(weekdays):
    return "".join(DAY_LETTERS.get(w.name.value, "") for w in weekdays)


def taught_with_partners(db, term_id, course_id):
    from models import TermTaughtWithGroup, TermTaughtWithMember

    member = (
        db.query(TermTaughtWithMember)
        .join(TermTaughtWithGroup)
        .filter(TermTaughtWithGroup.term_id == term_id, TermTaughtWithMember.course_id == course_id)
        .first()
    )
    if not member:
        return []
    return [m.course for m in member.group.members if m.course_id != course_id]


# --- Parsing the registrar's draft workbook ---------------------------------

def _s(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def _date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date().isoformat()
    return str(v)


def _row_from_excel(vals, idx):
    crn = vals[4]
    row_key = str(crn) if crn is not None else f"norow-{idx}"
    return {
        "row_key": row_key,
        "term_num": vals[1],
        "start_date": _date(vals[2]),
        "end_date": _date(vals[3]),
        "crn": crn,
        "subject": _s(vals[5]),
        "course_number": vals[6],
        "section": vals[7] or 1,
        "course_title": _s(vals[8]),
        "type": _s(vals[9]),
        "inst_method": _s(vals[10]),
        "instructor": _s(vals[11]),
        "secondary_instructor": _s(vals[12]),
        "hours": vals[13],
        "enrollment_max": vals[14],
        "waitlist_cap": vals[15],
        "begin": vals[16],
        "end": vals[17],
        "days": _s(vals[18]),
        "bldg": _s(vals[19]),
        "rm": _s(vals[20]),
        "course_comments": _s(vals[21]),
        "prerequisite": _s(vals[22]),
        "fee_detail": _s(vals[23]),
        "fee_amount": _s(vals[24]),
        "sig_code": _s(vals[25]),
        "sig_required": _s(vals[26]),
    }


def parse_workbook(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        if name.strip().upper() == "COMPOSITE":
            continue
        ws = wb[name]
        rows = []
        for i, vals in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            if len(vals) <= 6 or vals[6] is None:
                continue
            rows.append(_row_from_excel(vals, i))
        sheets[name] = rows
    return sheets


# --- Building the "new schedule" side from the DB ---------------------------

def build_new_rows(db, term, department) -> list:
    from models import ScheduleEntry, Course

    entries = (
        db.query(ScheduleEntry)
        .join(Course, ScheduleEntry.course_id == Course.id)
        .filter(
            Course.dept_code == department,
            ScheduleEntry.term_id == term.id,
            ScheduleEntry.schedule_table_id.isnot(None),
            ScheduleEntry.room_id.isnot(None),
        )
        .all()
    )

    rows = []
    for entry in entries:
        if not entry.time_slots:
            continue
        course = entry.course
        table = entry.schedule_table
        slots = sorted(entry.time_slots, key=lambda ts: ts.display_order)
        days = entry_days(entry, table)
        instructor = f"{entry.faculty.last_name}, {entry.faculty.first_name}" if entry.faculty else None

        partners = taught_with_partners(db, term.id, course.id)
        course_comments = None
        taughtwith_ref = None
        if partners:
            names = " and ".join(f"{p.dept_code} {p.course_number} {p.course_name}" for p in partners)
            course_comments = f"Taught with CRN [TBD] {names}"
            taughtwith_ref = (partners[0].dept_code, partners[0].course_number)

        rows.append({
            "row_key": f"add:{course.id}:{entry.section}",
            "course_id": course.id,
            "term_num": None,
            "start_date": None,
            "end_date": None,
            "crn": None,
            "subject": course.dept_code,
            "course_number": course.course_number,
            "section": entry.section,
            "course_title": course.course_name,
            "type": None,
            "inst_method": None,
            "instructor": instructor,
            "secondary_instructor": None,
            "hours": _credit_hours(course.course_number),
            "enrollment_max": None,  # resolved in diff()
            "waitlist_cap": None,
            "begin": time_to_hhmm(slots[0].start_time),
            "end": time_to_hhmm(slots[-1].end_time),
            "days": days_string(days),
            "bldg": entry.room.building_code,
            "rm": entry.room.room_number,
            "course_comments": course_comments,
            "taughtwith_ref": taughtwith_ref,
            "prerequisite": None,
            "fee_detail": None,
            "fee_amount": None,
            "sig_code": None,
            "sig_required": None,
            "_capacity": course.capacity,
        })
    return rows


# --- Diffing ------------------------------------------------------------

def _match_key(row):
    return (row.get("subject"), row.get("course_number"), row.get("section"))


def _room_of(row):
    return (str(row.get("bldg") or "").strip(), str(row.get("rm") or "").strip())


def diff(old_rows: list, new_rows: list, overrides: dict) -> list:
    old_by_key = {}
    for r in old_rows:
        old_by_key.setdefault(_match_key(r), r)

    new_by_key = {}
    for r in new_rows:
        new_by_key.setdefault(_match_key(r), r)

    sample_old = old_rows[0] if old_rows else None

    computed = []

    for key, old in old_by_key.items():
        new = new_by_key.get(key)
        if new is None:
            computed.append({
                "row_key": old["row_key"],
                "status": "delete",
                "changed_fields": [],
                "values": {k: v for k, v in old.items() if k in FIELD_ORDER or k == "row_key"},
                "original_enrollment_max": old.get("enrollment_max"),
            })
            continue

        row_key = old["row_key"]
        original_max = old.get("enrollment_max")
        effective_max = overrides.get(row_key, original_max)

        changed_fields = []
        if old.get("days") != new.get("days"):
            changed_fields.append("days")
        if old.get("begin") != new.get("begin"):
            changed_fields.append("begin")
        if old.get("end") != new.get("end"):
            changed_fields.append("end")
        if _room_of(old) != _room_of(new):
            changed_fields.append("room")
        if str(old.get("instructor") or "").strip().lower() != str(new.get("instructor") or "").strip().lower():
            changed_fields.append("instructor")
        if extract_taughtwith_ref(old.get("course_comments")) != new.get("taughtwith_ref"):
            changed_fields.append("course_comment")
        if effective_max != original_max:
            changed_fields.append("enrollment_max")

        merged = {k: v for k, v in old.items() if k in FIELD_ORDER or k == "row_key"}
        merged.update({
            "days": new.get("days"),
            "begin": new.get("begin"),
            "end": new.get("end"),
            "bldg": new.get("bldg"),
            "rm": new.get("rm"),
            "instructor": new.get("instructor"),
            "course_comments": new.get("course_comments") if new.get("course_comments") is not None else old.get("course_comments"),
            "enrollment_max": effective_max,
        })

        computed.append({
            "row_key": row_key,
            "status": "changed" if changed_fields else "keep",
            "changed_fields": changed_fields,
            "values": merged,
            "original_enrollment_max": original_max,
        })

    add_rows = [new for key, new in new_by_key.items() if key not in old_by_key]
    add_rows.sort(key=lambda r: (r.get("subject") or "", r.get("course_number") or 0, r.get("section") or 0))

    for new in add_rows:
        row_key = new["row_key"]
        default_max = new.get("_capacity")
        effective_max = overrides.get(row_key, default_max)
        merged = {
            "row_key": row_key,
            "term_num": sample_old["term_num"] if sample_old else None,
            "start_date": sample_old["start_date"] if sample_old else None,
            "end_date": sample_old["end_date"] if sample_old else None,
            "crn": None,
            "subject": new["subject"],
            "course_number": new["course_number"],
            "section": new["section"],
            "course_title": new["course_title"],
            "type": None,
            "inst_method": None,
            "instructor": new["instructor"],
            "secondary_instructor": None,
            "hours": new["hours"],
            "enrollment_max": effective_max,
            "waitlist_cap": None,
            "begin": new["begin"],
            "end": new["end"],
            "days": new["days"],
            "bldg": new["bldg"],
            "rm": new["rm"],
            "course_comments": new["course_comments"],
            "prerequisite": None,
            "fee_detail": None,
            "fee_amount": None,
            "sig_code": None,
            "sig_required": None,
        }
        computed.append({
            "row_key": row_key,
            "status": "add",
            "changed_fields": [],
            "values": merged,
            "original_enrollment_max": default_max,
        })

    return computed


# --- Excel export ------------------------------------------------------

def to_excel(computed_rows: list, department: str, term) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = department[:31]

    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")

    title = f"{term.semester.name.value.upper()} {term.year}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)

    for col, hdr in enumerate(HEADER, start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")

    row_idx = 3
    for row in computed_rows:
        status = row["status"]
        values = row["values"]
        changed_fields = row.get("changed_fields", [])

        ws.cell(row=row_idx, column=1, value=STATUS_LABEL[status])
        for col, field in enumerate(FIELD_ORDER, start=2):
            ws.cell(row=row_idx, column=col, value=values.get(field))

        if status == "keep":
            ws.cell(row=row_idx, column=1).fill = green
        elif status == "changed":
            ws.cell(row=row_idx, column=1).fill = yellow
            changed_cols = set()
            for cf in changed_fields:
                changed_cols.update(CHANGED_FIELD_COLUMNS.get(cf, []))
            for col, field in enumerate(FIELD_ORDER, start=2):
                if field in changed_cols:
                    ws.cell(row=row_idx, column=col).fill = yellow
        elif status == "delete":
            for col in range(1, len(HEADER) + 1):
                ws.cell(row=row_idx, column=col).fill = red
        elif status == "add":
            for col in range(1, len(HEADER) + 1):
                ws.cell(row=row_idx, column=col).fill = green

        row_idx += 1

    ws.column_dimensions["A"].width = 22
    for col in range(2, len(HEADER) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
